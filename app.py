import os
from io import BytesIO

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# =========================
# Excel読込設定
# =========================

TEMPLATE_FILE = "templates.xlsx"


# =========================
# テンプレートExcel読込
# =========================

def load_templates_from_excel(file_path: str = TEMPLATE_FILE) -> dict:
    """
    templates.xlsx から測定テンプレートを読み込む。

    推奨シート名：
    ・templates

    必要な列：
    ・テンプレート名
    ・No
    ・測定項目

    ※ templates シートがない場合は、1枚目のシートを読み込む。
    """

    if not os.path.exists(file_path):
        st.error(
            f"{file_path} が見つかりません。"
            "app.py と同じフォルダに templates.xlsx を置いてください。"
        )
        st.stop()

    try:
        try:
            template_df = pd.read_excel(file_path, sheet_name="templates")
        except ValueError:
            template_df = pd.read_excel(file_path, sheet_name=0)
    except Exception as e:
        st.error(
            "templates.xlsx の読み込みに失敗しました。"
            "1枚目のシート、または templates シートに"
            "「テンプレート名」「No」「測定項目」の列があるか確認してください。"
        )
        st.exception(e)
        st.stop()

    required_columns = ["テンプレート名", "No", "測定項目"]

    missing_columns = [
        col for col in required_columns
        if col not in template_df.columns
    ]

    if missing_columns:
        st.error(
            "templates.xlsx に必要な列がありません："
            + "、".join(missing_columns)
        )
        st.stop()

    template_df = template_df[required_columns].copy()

    template_df = template_df.dropna(
        subset=["テンプレート名", "No", "測定項目"]
    )

    template_df["テンプレート名"] = (
        template_df["テンプレート名"].astype(str).str.strip()
    )
    template_df["No"] = template_df["No"].astype(str).str.strip()
    template_df["測定項目"] = (
        template_df["測定項目"].astype(str).str.strip()
    )

    template_df = template_df[
        (template_df["テンプレート名"] != "") &
        (template_df["No"] != "") &
        (template_df["測定項目"] != "")
    ]

    templates = {}

    for template_name, group in template_df.groupby(
        "テンプレート名",
        sort=False
    ):
        rows = []
        for _, row in group.iterrows():
            rows.append({
                "No": row["No"],
                "測定項目": row["測定項目"],
            })
        templates[template_name] = rows

    if not templates:
        st.error("templates.xlsx に有効な測定項目がありません。")
        st.stop()

    return templates


# =========================
# データ作成
# =========================

def build_initial_df(template_name: str, templates: dict) -> pd.DataFrame:
    rows = []

    for item in templates[template_name]:
        rows.append({
            "No": item["No"],
            "測定項目": item["測定項目"],
            "測定": False,
            "写真": False,
        })

    return pd.DataFrame(rows)


# =========================
# Excel出力
# =========================

def create_excel(df: pd.DataFrame, template_name: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "チェック表"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    ng_fill = PatternFill("solid", fgColor="FCE4D6")

    ws["A1"] = "作工 測り漏れチェック表"
    ws["A1"].font = Font(size=16, bold=True)

    ws["A3"] = "テンプレート"
    ws["B3"] = template_name
    ws["A3"].font = Font(bold=True)
    ws["A3"].border = border
    ws["B3"].border = border

    table_start_row = 6
    columns = ["No", "測定項目", "測定", "写真"]

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(table_start_row, col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, row in enumerate(
        df.itertuples(index=False),
        start=table_start_row + 1
    ):
        values = [
            row.No,
            row.測定項目,
            "✓" if row.測定 else "",
            "✓" if row.写真 else "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            if col_idx in [3, 4]:
                cell.fill = ok_fill if value == "✓" else ng_fill

        ws.cell(row_idx, 2).alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

    ws2 = wb.create_sheet("未完了一覧")
    ws2["A1"] = "未完了一覧"
    ws2["A1"].font = Font(size=16, bold=True)

    incomplete_rows = []

    for _, row in df.iterrows():
        if not row["測定"]:
            incomplete_rows.append([
                row["No"],
                row["測定項目"],
                "測定"
            ])

        if not row["写真"]:
            incomplete_rows.append([
                row["No"],
                row["測定項目"],
                "写真"
            ])

    headers = ["No", "測定項目", "未完了項目"]

    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws2.cell(3, col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    if incomplete_rows:
        for row_idx, values in enumerate(incomplete_rows, start=4):
            for col_idx, value in enumerate(values, start=1):
                cell = ws2.cell(row_idx, col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            ws2.cell(row_idx, 2).alignment = Alignment(
                horizontal="left",
                vertical="center"
            )
    else:
        ws2["A4"] = "未完了項目はありません。"

    for sheet in wb.worksheets:
        widths = {
            "A": 8,
            "B": 30,
            "C": 12,
            "D": 12,
        }

        for col, width in widths.items():
            sheet.column_dimensions[col].width = width

        for row_num in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row_num].height = 28

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


# =========================
# Streamlit設定
# =========================

st.set_page_config(
    page_title="作工 測り漏れチェック",
    layout="centered"
)


# =========================
# スマホ向けCSS
# =========================

st.markdown(
    """
    <style>
    .block-container {
        padding-top: 1rem;
        padding-left: 0.8rem;
        padding-right: 0.8rem;
        max-width: 700px;
    }

    div[data-testid="stMetric"] {
        background-color: #f7f7f7;
        padding: 8px;
        border-radius: 8px;
    }

    .row-header {
        font-weight: bold;
        font-size: 16px;
        padding-top: 6px;
    }

    .item-name {
        font-size: 16px;
        padding-top: 6px;
    }

    div[data-testid="stCheckbox"] {
        padding-top: 2px;
    }

    div[data-testid="stCheckbox"] label {
        font-size: 18px;
        min-height: 36px;
    }
    </style>
    """,
    unsafe_allow_html=True
)


# =========================
# 初期化
# =========================

if "df" not in st.session_state:
    st.session_state.df = pd.DataFrame()

if "current_template" not in st.session_state:
    st.session_state.current_template = ""


# =========================
# 画面
# =========================

st.title("作工 測り漏れチェック")
st.caption("測定・写真のチェック専用アプリ")

templates = load_templates_from_excel()
template_names = list(templates.keys())

if st.session_state.current_template not in template_names:
    st.session_state.current_template = template_names[0]


# =========================
# テンプレート選択
# =========================

template_name = st.selectbox(
    "測定テンプレート",
    template_names,
    index=template_names.index(st.session_state.current_template)
)

if st.button("チェックリスト作成・リセット", use_container_width=True):
    st.session_state.df = build_initial_df(template_name, templates)
    st.session_state.current_template = template_name
    st.success("チェックリストを作成しました。")

if st.session_state.df.empty:
    st.info("測定テンプレートを選んで、チェックリストを作成してください。")
    st.stop()


# =========================
# 完了率
# =========================

df = st.session_state.df.copy()

total = len(df)
measure_done = int(df["測定"].sum())
photo_done = int(df["写真"].sum())

overall_rate = (measure_done + photo_done) / (total * 2) if total else 0

st.subheader(st.session_state.current_template)

st.write(f"全体完了率：{overall_rate:.0%}")
st.progress(overall_rate)

col1, col2 = st.columns(2)
col1.metric("測定", f"{measure_done} / {total}")
col2.metric("写真", f"{photo_done} / {total}")


# =========================
# 表示切替
# =========================

filter_col1, filter_col2 = st.columns(2)

with filter_col1:
    show_measure_none = st.checkbox("測定なし")

with filter_col2:
    show_photo_none = st.checkbox("写真なし")

display_df = df.copy()

if show_measure_none and show_photo_none:
    display_df = display_df[
        (display_df["測定"] == False) |
        (display_df["写真"] == False)
    ]
elif show_measure_none:
    display_df = display_df[display_df["測定"] == False]
elif show_photo_none:
    display_df = display_df[display_df["写真"] == False]


# =========================
# チェック表
# =========================

st.markdown("---")

header_cols = st.columns([0.8, 3.2, 1.1, 1.1])
header_cols[0].markdown("**No**")
header_cols[1].markdown("**測定項目**")
header_cols[2].markdown("**測定**")
header_cols[3].markdown("**写真**")

for idx, row in display_df.iterrows():
    cols = st.columns([0.8, 3.2, 1.1, 1.1])

    cols[0].markdown(
        f"<div class='row-header'>{row['No']}</div>",
        unsafe_allow_html=True
    )

    cols[1].markdown(
        f"<div class='item-name'>{row['測定項目']}</div>",
        unsafe_allow_html=True
    )

    measure_key = f"measure_{idx}"
    photo_key = f"photo_{idx}"

    new_measure = cols[2].checkbox(
        "",
        value=bool(st.session_state.df.at[idx, "測定"]),
        key=measure_key,
        label_visibility="collapsed"
    )

    new_photo = cols[3].checkbox(
        "",
        value=bool(st.session_state.df.at[idx, "写真"]),
        key=photo_key,
        label_visibility="collapsed"
    )

    st.session_state.df.at[idx, "測定"] = new_measure
    st.session_state.df.at[idx, "写真"] = new_photo

    st.markdown("---")


# =========================
# 未完了確認
# =========================

df_latest = st.session_state.df.copy()

measure_none = df_latest[df_latest["測定"] == False]
photo_none = df_latest[df_latest["写真"] == False]

st.subheader("未完了確認")

if measure_none.empty and photo_none.empty:
    st.success("未完了項目はありません。")
else:
    if not measure_none.empty:
        st.warning("測定なし")
        st.dataframe(
            measure_none[["No", "測定項目"]],
            use_container_width=True,
            hide_index=True
        )

    if not photo_none.empty:
        st.warning("写真なし")
        st.dataframe(
            photo_none[["No", "測定項目"]],
            use_container_width=True,
            hide_index=True
        )


# =========================
# Excel出力
# =========================

st.subheader("Excel出力")

excel_file = create_excel(
    df_latest,
    st.session_state.current_template
)

st.download_button(
    label="Excel出力",
    data=excel_file,
    file_name="作工_測り漏れチェック.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)